import os
import sys
import csv
import openpyxl
from flask import Flask, render_template, request, send_file, redirect, url_for, jsonify
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from collections import defaultdict
import pandas as pd
from pymongo import MongoClient
from pymongo.errors import ServerSelectionTimeoutError, ConnectionFailure
from dotenv import load_dotenv
import io
from bson.binary import Binary
from bson.objectid import ObjectId
import logging
import time
import socket

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv()

# Get the base directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, 
    static_folder=os.path.join(BASE_DIR, 'static'),
    static_url_path='/static',
    template_folder=os.path.join(BASE_DIR, 'templates')
)

# Add an upload size limit
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5 MB

# Path for uploaded files - use /tmp for Vercel
if os.environ.get('VERCEL'):
    UPLOAD_FOLDER = '/tmp/uploads'
else:
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

logger.info(f"Base directory: {BASE_DIR}")
logger.info(f"Template folder: {app.template_folder}")
logger.info(f"Static folder: {app.static_folder}")
logger.info(f"Upload folder: {UPLOAD_FOLDER}")
logger.info(f"Environment: {'VERCEL' if os.environ.get('VERCEL') else 'LOCAL'}")

# MongoDB connection - HARDCODED URI
MONGODB_URI = os.getenv('MONGODB_URI') or 'mongodb+srv://feedbackprocessing:surya1123@feedbackprocessing.cudqcoj.mongodb.net/?appName=feedbackprocessing&retryWrites=true&w=majority'

client = None
db = None
files_collection = None

logger.info(f"MongoDB URI set: {bool(MONGODB_URI)}")
logger.info(f"MongoDB URI preview: {MONGODB_URI[:80]}...")

def test_network():
    """Test basic network connectivity"""
    try:
        logger.info("Testing network connectivity to MongoDB...")
        sock = socket.create_connection(("feedbackprocessing.cudqcoj.mongodb.net", 27017), timeout=5)
        sock.close()
        logger.info("✓ Network connectivity OK")
        return True
    except Exception as e:
        logger.warning(f"⚠ Network test failed: {e}")
        return False

def connect_to_mongodb(retry_count=0, max_retries=3):
    """Connect to MongoDB with retry logic"""
    global client, db, files_collection
    
    if not MONGODB_URI:
        logger.error("✗ MONGODB_URI not available")
        return False
    
    try:
        logger.info(f"Attempting MongoDB connection (attempt {retry_count + 1}/{max_retries})...")
        
        # Test network first
        test_network()
        
        client = MongoClient(
            MONGODB_URI,
            serverSelectionTimeoutMS=30000,
            connectTimeoutMS=30000,
            socketTimeoutMS=30000,
            retryWrites=True,
            maxPoolSize=5,
            minPoolSize=1,
            ssl=True,
            tlsAllowInvalidCertificates=False,
            authSource='admin',
            authMechanism='SCRAM-SHA-1',
            maxIdleTimeMS=45000,
            waitQueueTimeoutMS=10000
        )
        
        # Test connection with timeout
        logger.info("Testing MongoDB connection with admin ping...")
        client.admin.command('ping')
        logger.info("✓ MongoDB ping successful!")
        
        # Access database
        db = client['feedback_processing']
        files_collection = db['processed_files']
        
        # Test collection
        logger.info("Testing collection access...")
        count = files_collection.count_documents({})
        logger.info(f"✓ MongoDB connected successfully! Found {count} files in database.")
        
        return True
        
    except (ServerSelectionTimeoutError, ConnectionFailure, TimeoutError) as e:
        logger.error(f"✗ MongoDB connection error (attempt {retry_count + 1}): {type(e).__name__}: {str(e)[:100]}")
        client = None
        db = None
        files_collection = None
        
        # Retry logic
        if retry_count < max_retries - 1:
            wait_time = 2 ** retry_count  # Exponential backoff
            logger.info(f"Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
            return connect_to_mongodb(retry_count + 1, max_retries)
        
        return False
        
    except Exception as e:
        logger.error(f"✗ Unexpected MongoDB error: {type(e).__name__}: {str(e)}", exc_info=True)
        client = None
        db = None
        files_collection = None
        return False

# Try to connect on startup
logger.info("Starting MongoDB connection attempt on app startup...")
try:
    connect_to_mongodb()
except Exception as e:
    logger.error(f"✗ Failed to connect to MongoDB on startup: {e}")

# Helper function to get a fresh connection if needed
def ensure_mongo_connection():
    """Ensure MongoDB connection is active, reconnect if needed"""
    global client, db, files_collection
    
    if files_collection is None:
        logger.warning("MongoDB not connected, attempting to reconnect...")
        return connect_to_mongodb()
    
    try:
        # Test the connection
        db.command('ping')
        return True
    except Exception as e:
        logger.warning(f"MongoDB connection lost, reconnecting... Error: {type(e).__name__}")
        return connect_to_mongodb()


def read_csv_headers(file_path):
    """Read CSV file and return all rows"""
    with open(file_path, mode='r', newline='', encoding='utf-8-sig') as file:
        csv_reader = csv.reader(file)
        rows = list(csv_reader)
    return rows


def create_header_dict(headers):
    """Create dictionary with header count and usage tracking"""
    header_count = defaultdict(int)
    for header in headers:
        header_count[header] += 1

    header_info = {}
    for header, count in header_count.items():
        header_info[header] = {
            'count': count,
            'used': 0
        }
    return header_info


def write_to_excel(rows, file_path):
    """Write data to an Excel file"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(file_path)


def process_worksheet(df, sheet, start_row):
    """Process worksheet to calculate totals and averages"""
    columns_with_dot = [col for col in df.columns if '.' in col]
    prefix_groups = {}

    for col in columns_with_dot:
        prefix = col.split('.')[0]
        if prefix not in prefix_groups:
            prefix_groups[prefix] = []
        prefix_groups[prefix].append(col)

    last_row = start_row + len(df) + 1

    for prefix, cols in prefix_groups.items():
        total_sums = df[cols].sum()
        for col in cols:
            col_index = df.columns.get_loc(col) + 2
            total_sum = total_sums[col]
            sheet.cell(row=last_row, column=col_index, value=total_sum)

            num_values = df[col].notna().sum()
            average_value = total_sum / num_values if num_values > 0 else 0
            average_percentage = round((average_value / 5) * 100, 2)
            sheet.cell(row=last_row + 1, column=col_index, value=average_percentage)

        avg_of_avgs = df[cols].mean(axis=1).mean() / 5 * 100
        first_col_in_group = df.columns.get_loc(cols[0]) + 2
        sheet.cell(row=last_row + 2, column=first_col_in_group, value=round(avg_of_avgs, 2))


def add_slno_column(sheet, df):
    """Add serial number column"""
    sheet.insert_cols(1)
    sheet.cell(row=1, column=1, value='SlNo')
    for row_num in range(2, len(df) + 2):
        sheet.cell(row=row_num, column=1, value=row_num - 1)


def apply_borders(sheet, start_row, end_row, start_col, end_col):
    """Apply borders to a range of cells"""
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border


def create_summary_sheet(workbook, df, file_path, branch_name):
    """Create summary sheet for each branch"""
    df = df.copy()

    if f"{branch_name} - Summary" not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet(f"{branch_name} - Summary")
    else:
        summary_sheet = workbook[f"{branch_name} - Summary"]

    file_name = os.path.basename(file_path)
    file_name_without_extension = os.path.splitext(file_name)[0]
    parts = file_name_without_extension.split('_')
    
    if len(parts) >= 3:
        academic_year, department, semester_info = parts[0], parts[1], parts[2]
    else:
        academic_year = parts[0] if len(parts) > 0 else 'Unknown Year'
        department = parts[1] if len(parts) > 1 else 'Unknown Department'
        semester_info = parts[2] if len(parts) > 2 else 'Unknown Semester'

    class_year, semester_num = semester_info.split('-') if '-' in semester_info else ('Unknown Class', 'Unknown Semester')
    class_year_display = f"{class_year} B.Tech"
    semester_display = f"{semester_num} Semester"

    if 'Timestamp' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Timestamp'].str.extract(r'(\d{4}/\d{2}/\d{2})')[0], errors='coerce')
        oldest_date = df['Timestamp'].min()
        oldest_date_str = oldest_date.strftime('%d-%m-%Y') if pd.notnull(oldest_date) else 'N/A'
    else:
        oldest_date_str = 'N/A'

    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    summary_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    summary_sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    summary_sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)

    summary_sheet.cell(row=1, column=1).value = "NARASARAOPETA ENGINEERING COLLEGE (AUTONOMOUS) - NARASARAOPET"
    summary_sheet.cell(row=2, column=1).value = f"STUDENT FEEDBACK SUMMARY - ACADEMIC YEAR - {academic_year}"
    summary_sheet.cell(row=3, column=1).value = f"DEPARTMENT OF {department}"
    summary_sheet.cell(row=4, column=1).value = f"CLASS: {class_year_display} {semester_display}, Section - {branch_name}, Date of Feedback: {oldest_date_str}"

    for row in range(1, 5):
        cell = summary_sheet.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    apply_borders(summary_sheet, start_row=1, end_row=4, start_col=1, end_col=11)

    row_num = 5
    averages_dict = {}
    
    for col in df.columns:
        if col == 'Timestamp':
            continue
        if '.' in col:
            subject_name = col.split('.')[0]
            avg_value = pd.to_numeric(df[col], errors='coerce').mean()
            if pd.notnull(avg_value):
                avg_percentage = (avg_value * 100) / 5
                if subject_name not in averages_dict:
                    averages_dict[subject_name] = []
                averages_dict[subject_name].append(round(avg_percentage, 2))

    bold_font = Font(bold=True)
    for subject_name, avg_values in averages_dict.items():
        summary_sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        summary_sheet.merge_cells(start_row=row_num, start_column=10, end_row=row_num, end_column=11)

        feedback_cell = summary_sheet.cell(row=row_num, column=1, value=f"Feedback for {subject_name}")
        feedback_cell.font = bold_font
        feedback_cell.alignment = Alignment(horizontal='left')

        responses_cell = summary_sheet.cell(row=row_num, column=10, value=f"No. of Responses {df.shape[0]}")
        responses_cell.font = bold_font
        responses_cell.alignment = Alignment(horizontal='right')

        apply_borders(summary_sheet, start_row=row_num, end_row=row_num, start_col=1, end_col=11)
        row_num += 1

        questions = ["Questions", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10"]
        for col_num, question in enumerate(questions, start=1):
            cell = summary_sheet.cell(row=row_num, column=col_num, value=question)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        apply_borders(summary_sheet, start_row=row_num, end_row=row_num, start_col=1, end_col=11)
        row_num += 1

        summary_sheet.cell(row=row_num, column=1, value="Percentage").font = bold_font
        for col_num, value in enumerate(avg_values, start=2):
            summary_sheet.cell(row=row_num, column=col_num, value=value)

        apply_borders(summary_sheet, start_row=row_num, end_row=row_num, start_col=1, end_col=11)
        
        avg_of_avg = sum(avg_values) / len(avg_values) if avg_values else 0
        summary_sheet.cell(row=row_num + 1, column=1, value="Feedback").font = bold_font
        summary_sheet.cell(row=row_num + 1, column=2, value=round(avg_of_avg, 2))
        apply_borders(summary_sheet, start_row=row_num + 1, end_row=row_num + 1, start_col=1, end_col=11)
        row_num += 2

    workbook.save(file_path)


def process_all_sheets(df, workbook, unique_column, file_path):
    """Process all sheets based on unique column values"""
    sheet = workbook.active
    add_slno_column(sheet, df)
    process_worksheet(df, sheet, start_row=1)

    if unique_column in df.columns:
        unique_values = df[unique_column].unique()
        for value in unique_values:
            if value not in workbook.sheetnames:
                new_sheet = workbook.create_sheet(title=str(value))
                filtered_data = df[df[unique_column] == value]
                add_slno_column(new_sheet, filtered_data)

                for col_num, column_title in enumerate(filtered_data.columns, start=2):
                    new_sheet.cell(row=1, column=col_num, value=column_title)

                for row_num, row_data in enumerate(filtered_data.values, start=2):
                    for col_num, cell_value in enumerate(row_data, start=2):
                        new_sheet.cell(row=row_num, column=col_num, value=cell_value)

                process_worksheet(filtered_data, new_sheet, start_row=1)
                create_summary_sheet(workbook, filtered_data, file_path, branch_name=str(value))


def create_comments_sheet(workbook, df):
    """Create Comments worksheet"""
    if 'Comments' not in workbook.sheetnames:
        comments_sheet = workbook.create_sheet('Comments')
        unique_sections = sorted(df['SECTION'].unique())
        
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        for i, section in enumerate(unique_sections):
            comments_sheet.merge_cells(start_row=1, start_column=i * 2 + 1, end_row=1, end_column=i * 2 + 2)
            comments_sheet.cell(row=1, column=i * 2 + 1).value = section
            comments_sheet.cell(row=2, column=i * 2 + 1).value = "Comments about the Department"
            comments_sheet.cell(row=2, column=i * 2 + 2).value = "Any Suggestions about the College"
            comments_sheet.cell(row=2, column=i * 2 + 1).font = Font(bold=True)
            comments_sheet.cell(row=2, column=i * 2 + 2).font = Font(bold=True)

        last_two_columns = df.columns[-2:]
        for i, section in enumerate(unique_sections):
            section_values = df[df['SECTION'] == section][last_two_columns].values
            for row_index, value in enumerate(section_values):
                comments_sheet.cell(row=row_index + 3, column=i * 2 + 1, value=value[0])
                comments_sheet.cell(row=row_index + 3, column=i * 2 + 2, value=value[1])

        for row in comments_sheet.iter_rows():
            for cell in row:
                cell.border = border


def save_file_to_mongodb(file_path, original_filename):
    """Save processed Excel file to MongoDB Atlas"""
    ensure_mongo_connection()
    
    if not files_collection:
        raise Exception("MongoDB collection not initialized - unable to connect to database")
    
    try:
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        filename = str(original_filename) if original_filename else os.path.basename(file_path)
        
        file_record = {
            'filename': filename,
            'file_data': Binary(file_data),
            'upload_date': pd.Timestamp.now(),
            'file_size': len(file_data)
        }
        
        result = files_collection.insert_one(file_record)
        logger.info(f"✓ File saved to MongoDB with ID: {result.inserted_id}")
        return str(result.inserted_id)
    except Exception as e:
        logger.error(f"✗ Error saving file to MongoDB: {e}", exc_info=True)
        raise


# Routes
@app.route('/')
def index():
    logger.info("Rendering index.html")
    return render_template('index.html')


@app.route('/test')
def test():
    """Test route to verify app is running and MongoDB is connected"""
    ensure_mongo_connection()
    mongo_status = "✓ connected" if files_collection else "✗ disconnected"
    return jsonify({
        'status': 'ok',
        'message': 'Flask app is running',
        'mongodb_status': mongo_status,
        'mongodb_uri_set': bool(MONGODB_URI),
        'environment': 'VERCEL' if os.environ.get('VERCEL') else 'LOCAL'
    }), 200


@app.route('/files')
def files_page():
    """Render files management page"""
    return render_template('files.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    ensure_mongo_connection()
    
    if not files_collection:
        logger.error("✗ MongoDB collection not available for upload")
        return jsonify({
            'success': False, 
            'error': 'Database connection failed. Please try again.'
        }), 503

    if request.content_length and request.content_length > app.config['MAX_CONTENT_LENGTH']:
        return jsonify({'success': False, 'error': 'File too large (max 5 MB)'}), 413

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        try:
            file.save(file_path)
            logger.info(f"File saved to {file_path}")

            rows = read_csv_headers(file_path)
            rows[0] = [header.strip() for header in rows[0]]

            header_info = create_header_dict(rows[0])

            generated_columns = []
            for header in rows[0]:
                if header_info[header]["count"] > 1:
                    header_info[header]["used"] += 1
                    new_column_name = f"{header}.{header_info[header]['used']}"
                    generated_columns.append(new_column_name)
                else:
                    generated_columns.append(header)

            rows[0] = generated_columns

            file_base, _ = os.path.splitext(file_path)
            excel_path = f"{file_base}_processed.xlsx"

            write_to_excel(rows, excel_path)
            logger.info(f"Excel file created at {excel_path}")

            workbook = load_workbook(excel_path)
            main_df = pd.read_excel(excel_path, sheet_name=0)

            main_df = main_df.copy()

            unique_column = 'SECTION'
            process_all_sheets(main_df, workbook, unique_column, excel_path)
            create_comments_sheet(workbook, main_df)

            sheets = workbook.worksheets
            if len(sheets) > 2:
                sorted_middle_sheets = sorted(sheets[1:-1], key=lambda ws: ws.title)
                workbook._sheets = [sheets[0]] + sorted_middle_sheets + [sheets[-1]]

            workbook.save(excel_path)
            logger.info(f"Excel file processed and saved")

            try:
                excel_filename = os.path.basename(excel_path)
                file_id = save_file_to_mongodb(excel_path, excel_filename)
                
                logger.info(f"✓ File saved to MongoDB with ID: {file_id}")
                
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    if os.path.exists(excel_path):
                        os.remove(excel_path)
                except Exception as delete_error:
                    logger.warning(f"Could not delete file: {delete_error}")
                
                return jsonify({'success': True, 'file_id': file_id}), 200
            except Exception as e:
                logger.error(f"✗ Error saving to MongoDB: {e}", exc_info=True)
                if os.path.exists(file_path):
                    os.remove(file_path)
                if os.path.exists(excel_path):
                    os.remove(excel_path)
                return jsonify({'success': False, 'error': f'Failed to save file: {str(e)}'}), 500
        except Exception as e:
            logger.error(f"✗ Upload error: {e}", exc_info=True)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except:
                    pass
            return jsonify({'success': False, 'error': str(e)}), 500
    
    return jsonify({'success': False, 'error': 'Unknown error'}), 500


@app.route('/download/<file_id>', methods=['GET'])
def download_file(file_id):
    """Download file from MongoDB"""
    ensure_mongo_connection()
    
    try:
        if not files_collection:
            return jsonify({'success': False, 'error': 'Database not connected'}), 503
        
        file_record = files_collection.find_one({'_id': ObjectId(file_id)})
        
        if not file_record:
            return jsonify({'success': False, 'error': 'File not found'}), 404
        
        file_data = file_record['file_data']
        filename = file_record['filename']
        
        return send_file(
            io.BytesIO(file_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"✗ Error downloading file: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/files', methods=['GET'])
def get_files():
    """Get all processed files from MongoDB"""
    ensure_mongo_connection()
    
    try:
        logger.info("Fetching files from MongoDB...")
        
        if not files_collection:
            return jsonify({
                'success': False, 
                'files': [], 
                'error': 'Database not connected'
            }), 503
        
        files = list(files_collection.find({}, {
            'filename': 1,
            'upload_date': 1,
            'file_size': 1,
            '_id': 1
        }).sort('upload_date', -1))
        
        logger.info(f"✓ Found {len(files)} files in MongoDB")
        
        result_files = []
        for file in files:
            try:
                file_dict = {
                    '_id': str(file.get('_id', '')),
                    'filename': str(file.get('filename', 'unknown_file.xlsx')),
                    'file_size': int(file.get('file_size', 0)),
                    'upload_date': file.get('upload_date', pd.Timestamp.now()).isoformat() if hasattr(file.get('upload_date'), 'isoformat') else str(file.get('upload_date', ''))
                }
                result_files.append(file_dict)
            except Exception as e:
                logger.warning(f"Error processing file record: {e}")
                continue
        
        return jsonify({'success': True, 'files': result_files}), 200
    except Exception as e:
        logger.error(f"✗ Error fetching files: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'files': [], 'error': str(e)}), 500


@app.route('/api/files/<file_id>', methods=['DELETE'])
def delete_file_api(file_id):
    """Delete a file from MongoDB"""
    ensure_mongo_connection()
    
    try:
        if not files_collection:
            return jsonify({'success': False, 'message': 'Database not connected'}), 503
        
        result = files_collection.delete_one({'_id': ObjectId(file_id)})
        
        if result.deleted_count > 0:
            logger.info(f"✓ File deleted: {file_id}")
            return jsonify({'success': True, 'message': 'File deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'File not found'}), 404
    except Exception as e:
        logger.error(f"✗ Error deleting file: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    ensure_mongo_connection()
    
    try:
        if not db:
            return jsonify({
                'status': 'unhealthy',
                'mongodb': 'not connected'
            }), 503
        
        db.command('ping')
        return jsonify({
            'status': 'healthy',
            'mongodb': 'connected',
            'timestamp': pd.Timestamp.now().isoformat()
        }), 200
    except Exception as e:
        logger.error(f"✗ Health check failed: {e}")
        return jsonify({'status': 'unhealthy', 'mongodb': 'disconnected'}), 503


@app.errorhandler(404)
def not_found(error):
    logger.warning(f"404 error: {error}")
    return jsonify({'success': False, 'error': 'Not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    logger.error(f"500 error: {error}")
    return jsonify({'success': False, 'error': 'Internal server error'}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
